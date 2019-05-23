from elasticsearch import Elasticsearch 
import openpyxl
import json
es=Elasticsearch([{'host':'localhost','port':9200}])

def json_file():
    with open('config.json') as json_data_file:
        data = json.load(json_data_file)
    return data

data = json_file()

theFile = openpyxl.load_workbook(data['read_data'])
allSheetNames = theFile.sheetnames

def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "L2 Jupiter Columns " or "L2 Jupiter Columns":
                return cell_name

def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    return letter

def get_all_values_by_cell_letter(letter):
    list1=[]
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
        list1.append(currentSheet[cell_name].value)
    return(list1)                

for sheet in allSheetNames:
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    letter = get_column_letter(specificCellLetter)
    wordFrequency=get_all_values_by_cell_letter(letter)
    index1=(sheet.lower()).replace(" ","_")
    # print(index1)
    # print('+++++')
    # print wordFrequency
    dict1={}
    for key in wordFrequency:
        value =''
        dict1[key] = value
        # print(dict1)
        #print(dict1)
    dict1 = {k:v for k,v in dict1.items() if k !=  u'Column Name' and k !=  u'L2 Jupiter Columns ' and k != u'L2 Jupiter Columns'}
    # dict1 = {k:v for k,v in dict1.items() if k !=  u'L2 Jupiter Columns '}

    # print(dict1)
    
    # print(dict1)
    res=es.index(index=index1,doc_type='jdbc',body=dict1)